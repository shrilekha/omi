"""
Shared Excel <-> benchmarks-table logic, used by both the CLI scripts
(export_benchmarks_to_excel.py / import_benchmarks_from_excel.py) and the
/admin/benchmarks/export and /admin/benchmarks/import routes in app.py — one
implementation, two ways to run it (terminal or the admin web page).
"""
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from db import list_benchmarks, upsert_benchmark
from constants import SECTORS, GEOS, REVENUE_BANDS, METRICS

HEADER = [
    'Sector', 'Geo', 'Revenue Band', 'Benchmark Value', 'Sample Size',
    'Source', 'Effective Date', 'Created By',
]

HEADER_FILL = PatternFill(start_color='0F1923', end_color='0F1923', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')

SECTOR_BY_LABEL = {label: id_ for id_, label in SECTORS}
GEO_BY_LABEL = {label: id_ for id_, label in GEOS}
BAND_BY_LABEL = {label: id_ for id_, label in REVENUE_BANDS}
METRIC_IDS = dict(METRICS)


def sheet_title(metric_id, metric_label):
    """'{metric_id} — {label}', truncated to Excel's 31-char sheet name limit.
    metric_id always survives intact (it's short and comes first) — parse_metric_id
    reads it back out via title.split(' — ', 1)[0], so truncating the label never
    breaks the round trip. Invalid sheet-name characters are stripped from the
    label half only, since metric ids never contain them."""
    safe_label = re.sub(r'[\\/*?\[\]:]', '-', metric_label)
    return f'{metric_id} — {safe_label}'[:31]


def parse_metric_id(title):
    """Inverse of sheet_title(): the metric id is the part before ' — '."""
    return title.split(' — ', 1)[0]


def _build_sheet(ws, metric_id):
    ws.append(HEADER)
    for col in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    existing = {
        (r['sector'], r['geo'], r['revenue_band']): r
        for r in list_benchmarks(metric=metric_id)
    }

    for sector_id, sector_label in SECTORS:
        for geo_id, geo_label in GEOS:
            for band_id, band_label in REVENUE_BANDS:
                r = existing.get((sector_id, geo_id, band_id))
                ws.append([
                    sector_label, geo_label, band_label,
                    r['benchmark_value'] if r else '',
                    (r['sample_size'] if r and r['sample_size'] is not None else ''),
                    r['source'] if r else '',
                    r['effective_date'] if r else '',
                    r['created_by'] if r else '',
                ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    widths = [26, 20, 24, 16, 12, 26, 16, 18]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'


def build_workbook():
    """The full sector x geo x revenue-band grid, one sheet per metric,
    pre-filled with every existing value and a blank editable row for every
    combination that doesn't have one yet."""
    wb = Workbook()
    for i, (metric_id, metric_label) in enumerate(METRICS):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sheet_title(metric_id, metric_label)
        _build_sheet(ws, metric_id)
    return wb


def _format_date(value):
    if value in (None, ''):
        return None
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value)


def import_workbook(wb):
    """Upserts every non-blank row of an already-loaded Workbook into the
    benchmarks table, matched by natural key (metric, sector, geo, revenue_band).
    Safe to run more than once — re-importing the same file just updates the
    same rows again rather than duplicating them. Returns a summary dict."""
    inserted = updated = skipped = 0
    errors = []

    for ws in wb.worksheets:
        metric_id = parse_metric_id(ws.title)
        if metric_id not in METRIC_IDS:
            errors.append(f"sheet '{ws.title}': unrecognized metric id '{metric_id}' — skipping sheet")
            continue

        existing = {
            (r['sector'], r['geo'], r['revenue_band']): r['id']
            for r in list_benchmarks(metric=metric_id)
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            padded = list(row) + [None] * (8 - len(row))
            (sector_label, geo_label, band_label, value,
             sample_size, source, eff_date, created_by) = padded[:8]
            ref = f"'{ws.title}' row {row_idx}"

            if value in (None, ''):
                skipped += 1
                continue

            sector_id = SECTOR_BY_LABEL.get(sector_label)
            geo_id = GEO_BY_LABEL.get(geo_label)
            band_id = BAND_BY_LABEL.get(band_label)
            if not (sector_id and geo_id and band_id):
                errors.append(
                    f'{ref}: unrecognized sector/geo/revenue-band label '
                    f'({sector_label!r}, {geo_label!r}, {band_label!r})'
                )
                continue

            try:
                benchmark_value = float(value)
            except (TypeError, ValueError):
                errors.append(f'{ref}: benchmark value {value!r} is not numeric')
                continue

            sample_size_int = None
            if sample_size not in (None, ''):
                try:
                    sample_size_int = int(sample_size)
                except (TypeError, ValueError):
                    errors.append(f'{ref}: sample size {sample_size!r} is not an integer')
                    continue

            data = {
                'metric': metric_id, 'sector': sector_id, 'geo': geo_id, 'revenue_band': band_id,
                'benchmark_value': benchmark_value, 'sample_size': sample_size_int,
                'source': source or None,
                'effective_date': _format_date(eff_date),
                'created_by': created_by or None,
            }

            existing_id = existing.get((sector_id, geo_id, band_id))
            upsert_benchmark(data, existing_id)
            if existing_id:
                updated += 1
            else:
                inserted += 1

    return {'inserted': inserted, 'updated': updated, 'skipped': skipped, 'errors': errors}


def load_workbook_file(file_or_path):
    return load_workbook(file_or_path, data_only=True)
