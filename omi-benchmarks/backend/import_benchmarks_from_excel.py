"""
Imports a workbook produced by export_benchmarks_to_excel.py (or manually filled in
the same shape) back into the benchmarks table. Safe to re-run: rows are matched by
the table's natural key (metric, sector, geo, revenue_band) and upserted, so importing
the same file twice just updates the same rows again rather than duplicating them.

Usage:
    python import_benchmarks_from_excel.py [path/to/workbook.xlsx]

Default path: Benchmarks_Export.xlsx (next to this script)
"""
import os
import sys

from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

from openpyxl import load_workbook

from db import init_db, list_benchmarks, upsert_benchmark
from constants import SECTORS, GEOS, REVENUE_BANDS, METRICS

DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Benchmarks_Export.xlsx')

SECTOR_BY_LABEL = {label: id_ for id_, label in SECTORS}
GEO_BY_LABEL = {label: id_ for id_, label in GEOS}
BAND_BY_LABEL = {label: id_ for id_, label in REVENUE_BANDS}
METRIC_IDS = dict(METRICS)


def parse_metric_id(title):
    """Inverse of export's sheet_title(): the metric id is the part before ' — ',
    which is never truncated since it's short and written first."""
    return title.split(' — ', 1)[0]


def format_date(value):
    if value in (None, ''):
        return None
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    return str(value)


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(path):
        print(f'File not found: {path}')
        sys.exit(1)

    init_db()
    wb = load_workbook(path, data_only=True)

    inserted = updated = skipped = 0
    errors = []

    for ws in wb.worksheets:
        metric_id = parse_metric_id(ws.title)
        if metric_id not in METRIC_IDS:
            errors.append(f"sheet '{ws.title}': unrecognized metric id '{metric_id}' — skipping sheet")
            continue

        existing = {
            (r['sector'], r['geo'], r['revenue_band']): r['id']
            for r in list_benchmarks(metric=metric_id)
        }

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            padded = list(row) + [None] * (8 - len(row))
            (sector_label, geo_label, band_label, value,
             sample_size, source, eff_date, created_by) = padded[:8]
            ref = f"'{ws.title}' row {row_idx}"

            if value in (None, ''):
                skipped += 1
                continue

            sector_id = SECTOR_BY_LABEL.get(sector_label)
            geo_id = GEO_BY_LABEL.get(geo_label)
            band_id = BAND_BY_LABEL.get(band_label)
            if not (sector_id and geo_id and band_id):
                errors.append(
                    f'{ref}: unrecognized sector/geo/revenue-band label '
                    f'({sector_label!r}, {geo_label!r}, {band_label!r})'
                )
                continue

            try:
                benchmark_value = float(value)
            except (TypeError, ValueError):
                errors.append(f'{ref}: benchmark value {value!r} is not numeric')
                continue

            sample_size_int = None
            if sample_size not in (None, ''):
                try:
                    sample_size_int = int(sample_size)
                except (TypeError, ValueError):
                    errors.append(f'{ref}: sample size {sample_size!r} is not an integer')
                    continue

            data = {
                'metric': metric_id, 'sector': sector_id, 'geo': geo_id, 'revenue_band': band_id,
                'benchmark_value': benchmark_value, 'sample_size': sample_size_int,
                'source': source or None,
                'effective_date': format_date(eff_date),
                'created_by': created_by or None,
            }

            existing_id = existing.get((sector_id, geo_id, band_id))
            upsert_benchmark(data, existing_id)
            if existing_id:
                updated += 1
            else:
                inserted += 1

    print(f'Inserted: {inserted}  Updated: {updated}  Skipped (blank): {skipped}  Errors: {len(errors)}')
    for e in errors:
        print(f'  ! {e}')


if __name__ == '__main__':
    main()
