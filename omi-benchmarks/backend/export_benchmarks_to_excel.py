"""
Exports the omi-benchmarks grid — every sector x geo x revenue-band combination,
for every canonical metric — into a single Excel workbook for offline editing.
Run this to produce a fresh copy to hand to the content team, or to re-baseline
after edits have already been imported.

Usage:
    python export_benchmarks_to_excel.py

Output:
    Benchmarks_Export.xlsx (next to this script)
"""
import os
import re

from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from db import init_db, list_benchmarks
from constants import SECTORS, GEOS, REVENUE_BANDS, METRICS

OUTPUT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Benchmarks_Export.xlsx')

HEADER = [
    'Sector', 'Geo', 'Revenue Band', 'Benchmark Value', 'Sample Size',
    'Source', 'Effective Date', 'Created By',
]

HEADER_FILL = PatternFill(start_color='0F1923', end_color='0F1923', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')


def sheet_title(metric_id, metric_label):
    """'{metric_id} — {label}', truncated to Excel's 31-char sheet name limit.
    metric_id always survives intact (it's short and comes first) — the import
    script reads it back out via sheet_title.split(' — ', 1)[0], so truncating
    the label never breaks the round trip. Invalid sheet-name characters are
    stripped from the label half only, since metric ids never contain them."""
    safe_label = re.sub(r'[\\/*?\[\]:]', '-', metric_label)
    return f'{metric_id} — {safe_label}'[:31]


def build_sheet(ws, metric_id):
    ws.append(HEADER)
    for col in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    existing = {
        (r['sector'], r['geo'], r['revenue_band']): r
        for r in list_benchmarks(metric=metric_id)
    }

    for sector_id, sector_label in SECTORS:
        for geo_id, geo_label in GEOS:
            for band_id, band_label in REVENUE_BANDS:
                r = existing.get((sector_id, geo_id, band_id))
                ws.append([
                    sector_label, geo_label, band_label,
                    r['benchmark_value'] if r else '',
                    (r['sample_size'] if r and r['sample_size'] is not None else ''),
                    r['source'] if r else '',
                    r['effective_date'] if r else '',
                    r['created_by'] if r else '',
                ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    widths = [26, 20, 24, 16, 12, 26, 16, 18]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'


def main():
    init_db()
    wb = Workbook()
    row_total = 0
    for i, (metric_id, metric_label) in enumerate(METRICS):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = sheet_title(metric_id, metric_label)
        build_sheet(ws, metric_id)
        row_total += ws.max_row - 1

    wb.save(OUTPUT_XLSX)
    print(f'Wrote {len(METRICS)} metric sheets ({row_total} rows) to {OUTPUT_XLSX}')


if __name__ == '__main__':
    main()
