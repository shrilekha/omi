"""
Shared Excel <-> questions.py logic for omi-deepdive's /admin/questions page.

questions.py mixes Excel-worthy content (question text/hint/options, tool
question text) with structural fields that must NEVER come from Excel (
dimension id/name/icon/weight/na_allowed/na_label — weight has a hard
`assert sum(...) == 100` at module load, and dimension ids are referenced by
scoring.py and stored in submissions). So import here is a MERGE, not a
replace: every dimension in the current file keeps its own structural
fields, and only `questions` / `tool_question.text` (and OWNERSHIP_QUESTIONS'
text) are ever overwritten from a workbook. The dimension set itself is
fixed — Excel can't add or remove a dimension, that's a code change.

Only the two top-level data assignments (OWNERSHIP_QUESTIONS, DIMENSIONS)
are ever rewritten in questions.py — the header comment and MATURITY_BANDS/
the trailing weight-sum assert are left untouched. Import is all-or-nothing.
"""
import ast
import importlib
import os
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
QUESTIONS_PY = os.path.join(BACKEND_DIR, 'questions.py')

SHEET_OWNERSHIP = 'Ownership Questions'
SHEET_TOOL = 'Tool Questions'
SHEET_SCORED = 'Scored Questions'

HEADER_FILL = PatternFill(start_color='0F1923', end_color='0F1923', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')


def _load_questions_module():
    """Imports the live questions.py fresh each time (not cached), so a
    write made by a previous import in the same process is picked up by the
    next read without restarting the server."""
    if BACKEND_DIR not in sys.path:
        sys.path.insert(0, BACKEND_DIR)
    if 'questions' in sys.modules:
        return importlib.reload(sys.modules['questions'])
    return importlib.import_module('questions')


def load_current():
    """Returns (ownership_questions, dimensions) — live Python data from the
    current questions.py, not a re-parse."""
    mod = _load_questions_module()
    return mod.OWNERSHIP_QUESTIONS, mod.DIMENSIONS


# ─── Excel build (export) ───────────────────────────────────────────────────

def _write_header(ws, header):
    """Must be called BEFORE any data rows are appended — inserts at row 1."""
    ws.append(header)
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def _finalize_sheet(ws, widths):
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'


def build_workbook():
    ownership, dimensions = load_current()
    wb = Workbook()

    ws1 = wb.active
    ws1.title = SHEET_OWNERSHIP
    _write_header(ws1, ['Question ID', 'Question Text'])
    for q in ownership:
        ws1.append([q['id'], q['text']])
    _finalize_sheet(ws1, [22, 70])

    ws2 = wb.create_sheet(SHEET_TOOL)
    _write_header(
        ws2, ['Dimension ID', 'Dimension Name (reference only)', 'Tool Question ID (reference only)', 'Tool Question Text'],
    )
    for d in dimensions:
        tq = d['tool_question']
        ws2.append([d['id'], d['name'], tq['id'], tq['text']])
    _finalize_sheet(ws2, [16, 34, 26, 70])

    ws3 = wb.create_sheet(SHEET_SCORED)
    _write_header(
        ws3,
        ['Dimension ID', 'Dimension Name (reference only)', 'Question ID', 'Question Text', 'Hint',
         'Option 1 (score 1)', 'Option 2 (score 2)', 'Option 3 (score 3)',
         'Option 4 (score 4)', 'Option 5 (score 5)'],
    )
    for d in dimensions:
        for q in d['questions']:
            opts = q['options']
            ws3.append([
                d['id'], d['name'], q['id'], q['text'], q.get('hint', ''),
                opts[0], opts[1], opts[2], opts[3], opts[4],
            ])
    _finalize_sheet(ws3, [16, 34, 10, 50, 40, 34, 34, 34, 34, 34])

    return wb


# ─── Read rows out of an uploaded workbook (import) ─────────────────────────

def _row_dicts(ws):
    header = [c.value for c in ws[1]]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v in (None, '') for v in row):
            continue
        yield row_idx, dict(zip(header, row))


def read_workbook(wb, current_ownership, current_dimensions):
    """Returns (ownership_after, dimensions_after, errors). dimensions_after
    is a full MERGE of current_dimensions with Excel content — every
    structural field (name/icon/weight/na_allowed/na_label/id) comes from
    current_dimensions, only questions/tool_question text come from wb."""
    errors = []
    for required in (SHEET_OWNERSHIP, SHEET_TOOL, SHEET_SCORED):
        if required not in wb.sheetnames:
            errors.append(f"required sheet '{required}' not found in workbook")
    if errors:
        return None, None, errors

    current_dim_ids = {d['id'] for d in current_dimensions}
    current_ownership_ids = {q['id'] for q in current_ownership}

    # ── Ownership questions: same fixed 2 ids, text only ──
    ownership_text = {}
    for row_idx, d in _row_dicts(wb[SHEET_OWNERSHIP]):
        qid = str(d.get('Question ID') or '').strip()
        text = str(d.get('Question Text') or '').strip()
        ref = f"'{SHEET_OWNERSHIP}' row {row_idx}"
        if not qid:
            errors.append(f'{ref}: missing Question ID')
            continue
        if qid not in current_ownership_ids:
            errors.append(f"{ref}: '{qid}' is not an existing ownership question id "
                           f"— adding/renaming ownership questions isn't supported via Excel")
            continue
        if not text:
            errors.append(f"{ref} ({qid}): missing Question Text")
            continue
        ownership_text[qid] = text
    missing_ownership = current_ownership_ids - set(ownership_text)
    if missing_ownership:
        errors.append(f"'{SHEET_OWNERSHIP}': missing row(s) for {sorted(missing_ownership)}")

    # ── Tool questions: one row per dimension, text only (id is reference-only) ──
    tool_text = {}
    for row_idx, d in _row_dicts(wb[SHEET_TOOL]):
        dim_id = str(d.get('Dimension ID') or '').strip()
        text = str(d.get('Tool Question Text') or '').strip()
        ref = f"'{SHEET_TOOL}' row {row_idx}"
        if dim_id not in current_dim_ids:
            errors.append(f"{ref}: unrecognized Dimension ID '{dim_id}'")
            continue
        if not text:
            errors.append(f"{ref} ({dim_id}): missing Tool Question Text")
            continue
        tool_text[dim_id] = text
    missing_tool = current_dim_ids - set(tool_text)
    if missing_tool:
        errors.append(f"'{SHEET_TOOL}': missing row(s) for {sorted(missing_tool)}")

    # ── Scored questions: grouped by dimension, full question content ──
    scored_by_dim = {d['id']: [] for d in current_dimensions}
    for row_idx, d in _row_dicts(wb[SHEET_SCORED]):
        dim_id = str(d.get('Dimension ID') or '').strip()
        ref = f"'{SHEET_SCORED}' row {row_idx}"
        if dim_id not in current_dim_ids:
            errors.append(f"{ref}: unrecognized Dimension ID '{dim_id}'")
            continue
        qid = str(d.get('Question ID') or '').strip()
        text = str(d.get('Question Text') or '').strip()
        hint = str(d.get('Hint') or '').strip()
        if not qid:
            errors.append(f'{ref}: missing Question ID')
        if not text:
            errors.append(f'{ref} ({qid or "?"}): missing Question Text')

        options = []
        for score in range(1, 6):
            col = f'Option {score} (score {score})'
            val = str(d.get(col) or '').strip()
            if not val:
                errors.append(f'{ref} ({qid or "?"}): missing Option {score}')
            options.append(val)

        if qid and text and all(options):
            q = {'id': qid, 'text': text, 'options': options}
            if hint:
                q['hint'] = hint
            scored_by_dim[dim_id].append(q)

    for dim_id, questions in scored_by_dim.items():
        if not questions:
            errors.append(f"'{SHEET_SCORED}': no rows found for dimension '{dim_id}' "
                           f"— refusing to empty out its question list")
        ids = [q['id'] for q in questions]
        dupes = {i for i in ids if ids.count(i) > 1}
        if dupes:
            errors.append(f"'{SHEET_SCORED}'.{dim_id}: duplicate question ids {sorted(dupes)}")

    if errors:
        return None, None, errors

    ownership_after = [{**q, 'text': ownership_text[q['id']]} for q in current_ownership]
    dimensions_after = []
    for d in current_dimensions:
        merged = dict(d)
        merged['tool_question'] = {**d['tool_question'], 'text': tool_text[d['id']]}
        merged['questions'] = scored_by_dim[d['id']]
        dimensions_after.append(merged)

    assert sum(d['weight'] for d in dimensions_after) == 100, \
        'merge altered dimension weights — this should be impossible'

    return ownership_after, dimensions_after, errors


def read_workbook_file(path_or_file):
    current_ownership, current_dimensions = load_current()
    wb = load_workbook(path_or_file, data_only=True)
    return read_workbook(wb, current_ownership, current_dimensions)


# ─── Serialize back to Python source, matching questions.py's style ───────────

def _py_string(s):
    escaped = s.replace('\\', '\\\\').replace('"', '\\"')
    return f'"{escaped}"'


def _serialize_ownership(questions, indent=0):
    pad = '  ' * indent
    inner = pad + '    '
    if not questions:
        return '[]'
    items = []
    for q in questions:
        items.append(
            f'{pad}    {{\n'
            f'{inner}    "id": {_py_string(q["id"])},\n'
            f'{inner}    "text": {_py_string(q["text"])},\n'
            f'{pad}    }}'
        )
    return '[\n' + ',\n'.join(items) + '\n' + pad + ']'


def _serialize_question(q, indent):
    pad = '  ' * indent
    inner = pad + '  '
    lines = [pad + '{', inner + f'"id": {_py_string(q["id"])},', inner + f'"text": {_py_string(q["text"])},']
    if q.get('hint'):
        lines.append(inner + f'"hint": {_py_string(q["hint"])},')
    lines.append(inner + '"options": [')
    opt_pad = inner + '    '
    lines.append(',\n'.join(opt_pad + _py_string(o) for o in q['options']))
    lines.append(inner + '],')
    lines.append(pad + '},')
    return '\n'.join(lines)


def _serialize_dimension(d, indent=1):
    pad = '  ' * indent
    inner = pad + '  '
    tq = d['tool_question']
    lines = [pad + '{']
    lines.append(inner + f'"id": {_py_string(d["id"])},')
    lines.append(inner + f'"name": {_py_string(d["name"])},')
    lines.append(inner + f'"icon": {_py_string(d["icon"])},')
    lines.append(inner + f'"weight": {d["weight"]},')
    lines.append(inner + f'"na_allowed": {"True" if d.get("na_allowed") else "False"},')
    if d.get('na_allowed') and d.get('na_label'):
        lines.append(inner + f'"na_label": {_py_string(d["na_label"])},')
    lines.append(inner + '"tool_question": {')
    lines.append(inner + f'    "id": {_py_string(tq["id"])},')
    lines.append(inner + f'    "text": {_py_string(tq["text"])},')
    lines.append(inner + '},')
    lines.append(inner + '"questions": [')
    lines.append('\n'.join(_serialize_question(q, indent + 2) for q in d['questions']))
    lines.append(inner + '],')
    lines.append(pad + '},')
    return '\n'.join(lines)


def serialize_ownership(questions):
    return f'OWNERSHIP_QUESTIONS = {_serialize_ownership(questions)}'


def serialize_dimensions(dimensions):
    body = '\n'.join(_serialize_dimension(d, 1) for d in dimensions)
    return f'DIMENSIONS = [\n{body}\n]'


# ─── Locate DIMENSIONS / OWNERSHIP_QUESTIONS via ast, splice in new source ────

def _char_offset(lines_with_ends, lineno, col_offset):
    return sum(len(l) for l in lines_with_ends[:lineno - 1]) + col_offset


def _locate_assignment(source, var_name):
    tree = ast.parse(source)
    for node in tree.body:
        if (isinstance(node, ast.Assign) and len(node.targets) == 1
                and isinstance(node.targets[0], ast.Name) and node.targets[0].id == var_name):
            lines = source.splitlines(keepends=True)
            start = _char_offset(lines, node.lineno, node.col_offset)
            end = _char_offset(lines, node.end_lineno, node.end_col_offset)
            return start, end
    raise ValueError(f'{var_name} not found in {QUESTIONS_PY}')


def splice(source, replacements):
    spans = {name: _locate_assignment(source, name) for name in replacements}
    ordered = sorted(spans.items(), key=lambda kv: kv[1][0])
    out, cursor = [], 0
    for name, (start, end) in ordered:
        out.append(source[cursor:start])
        out.append(replacements[name])
        cursor = end
    out.append(source[cursor:])
    return ''.join(out)


# ─── Diff summary ───────────────────────────────────────────────────────────

def _dim_questions_by_id(dimensions):
    return {d['id']: {q['id']: q for q in d['questions']} for d in dimensions}


def diff_lines(ownership_before, ownership_after, dims_before, dims_after):
    lines = []

    before_own = {q['id']: q for q in ownership_before}
    after_own = {q['id']: q for q in ownership_after}
    changed_own = sorted(qid for qid in before_own if before_own[qid] != after_own.get(qid))
    if changed_own:
        lines.append(f'OWNERSHIP_QUESTIONS: text changed for {changed_own}')

    before_tool = {d['id']: d['tool_question']['text'] for d in dims_before}
    after_tool = {d['id']: d['tool_question']['text'] for d in dims_after}
    changed_tool = sorted(dim for dim in before_tool if before_tool[dim] != after_tool.get(dim))
    if changed_tool:
        lines.append(f'tool_question text changed for dimensions {changed_tool}')

    before_q, after_q = _dim_questions_by_id(dims_before), _dim_questions_by_id(dims_after)
    for dim_id in sorted(before_q):
        b, a = before_q[dim_id], after_q.get(dim_id, {})
        added = sorted(set(a) - set(b))
        removed = sorted(set(b) - set(a))
        changed = sorted(qid for qid in (set(a) & set(b)) if a[qid] != b[qid])
        if added or removed or changed:
            bits = f'DIMENSIONS.{dim_id}.questions: {len(b)} -> {len(a)} questions'
            if added:
                bits += f'  +{added}'
            if removed:
                bits += f'  -{removed}'
            if changed:
                bits += f'  ~{changed}'
            lines.append(bits)

    return lines


def apply_import(ownership_after, dimensions_after):
    """Writes the merged data into questions.py (backing up the previous
    version to questions.py.bak first) and returns human-readable diff
    lines against what was there before."""
    ownership_before, dims_before = load_current()

    with open(QUESTIONS_PY, 'r', encoding='utf-8') as f:
        source = f.read()

    backup_path = QUESTIONS_PY + '.bak'
    with open(backup_path, 'w', encoding='utf-8') as f:
        f.write(source)

    new_source = splice(source, {
        'OWNERSHIP_QUESTIONS': serialize_ownership(ownership_after),
        'DIMENSIONS': serialize_dimensions(dimensions_after),
    })

    with open(QUESTIONS_PY, 'w', encoding='utf-8') as f:
        f.write(new_source)

    return diff_lines(ownership_before, ownership_after, dims_before, dimensions_after)
