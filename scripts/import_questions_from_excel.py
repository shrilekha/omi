"""
Imports a workbook shaped like the one export_questions_to_excel.py produces back
into frontend/questions.js. Only the 5 data literals (txnVariants, compVariants,
appQuestions, infraQuestions, logQuestions) are touched — everything else in the
file (the header comment block, sectorArchetypeMap, resolveArchetype(),
getSections()) is left byte-for-byte untouched, since questions.js has real logic
in it, not just data.

Reads sheets 'Txn Observability', 'Compliance & Audit', and
'App, Infra & Log (shared)' — the same three the export script builds them from.
'All questions (combined)' is a derived convenience view for reviewers and is
intentionally NOT read here, so there's exactly one place edits can be made that
takes effect.

All-or-nothing: if any row fails validation, nothing is written — the run prints
every problem it found instead.

Usage:
    python scripts/import_questions_from_excel.py [path/to/workbook.xlsx]

Default path: OMI_Questions_Review.xlsx (project root, export script's own default)
"""
import os
import re
import sys

from openpyxl import load_workbook

from export_questions_to_excel import (
    QUESTIONS_JS, load_questions, _parse_value,
)

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
DEFAULT_XLSX = os.path.join(ROOT, 'OMI_Questions_Review.xlsx')

SHEET_TXN = 'Txn Observability'
SHEET_COMP = 'Compliance & Audit'
SHEET_SHARED = 'App, Infra & Log (shared)'

DOMAIN_TO_SHARED_VAR = {
    'Application Performance': 'appQuestions',
    'Infrastructure & Network': 'infraQuestions',
    'Log Management & Data Lake': 'logQuestions',
}


# ─── Read rows out of the workbook ──────────────────────────────────────────

def _row_dicts(ws):
    """Yield (row_number, {header: value}) for every non-empty data row."""
    header = [c.value for c in ws[1]]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v in (None, '') for v in row):
            continue
        yield row_idx, dict(zip(header, row))


def _question_from_row(row_idx, ref, d, errors):
    qid = str(d.get('Question ID') or '').strip()
    text = str(d.get('Question Text') or '').strip()
    hint = str(d.get('Hint') or '').strip()

    if not qid:
        errors.append(f'{ref} row {row_idx}: missing Question ID')
    if not text:
        errors.append(f'{ref} row {row_idx}: missing Question Text')

    options = []
    for score in range(1, 6):
        col = f'Option {score} (score {score})'
        val = str(d.get(col) or '').strip()
        if not val:
            errors.append(f'{ref} row {row_idx} ({qid or "?"}): missing Option {score}')
            continue
        options.append({'text': val, 'score': score})

    q = {'id': qid, 'text': text, 'options': options}
    if hint:
        q['hint'] = hint
    return q


def read_workbook(path):
    """Returns (txn_variants, comp_variants, app_q, infra_q, log_q, errors)."""
    wb = load_workbook(path, data_only=True)
    errors = []
    for required in (SHEET_TXN, SHEET_COMP, SHEET_SHARED):
        if required not in wb.sheetnames:
            errors.append(f"required sheet '{required}' not found in workbook")
    if errors:
        return None, None, None, None, None, errors

    txn_variants = {}
    for row_idx, d in _row_dicts(wb[SHEET_TXN]):
        variant = str(d.get('Sector / Country Variant') or '').strip()
        if not variant:
            errors.append(f"'{SHEET_TXN}' row {row_idx}: missing Sector / Country Variant")
            continue
        q = _question_from_row(row_idx, f"'{SHEET_TXN}'", d, errors)
        txn_variants.setdefault(variant, []).append(q)

    comp_variants = {}
    for row_idx, d in _row_dicts(wb[SHEET_COMP]):
        variant = str(d.get('Sector / Country Variant') or '').strip()
        if not variant:
            errors.append(f"'{SHEET_COMP}' row {row_idx}: missing Sector / Country Variant")
            continue
        q = _question_from_row(row_idx, f"'{SHEET_COMP}'", d, errors)
        comp_variants.setdefault(variant, []).append(q)

    app_q, infra_q, log_q = [], [], []
    target_by_var = {'appQuestions': app_q, 'infraQuestions': infra_q, 'logQuestions': log_q}
    for row_idx, d in _row_dicts(wb[SHEET_SHARED]):
        domain = str(d.get('Domain') or '').strip()
        var_name = DOMAIN_TO_SHARED_VAR.get(domain)
        if not var_name:
            errors.append(f"'{SHEET_SHARED}' row {row_idx}: unrecognized Domain '{domain}'")
            continue
        q = _question_from_row(row_idx, f"'{SHEET_SHARED}'", d, errors)
        target_by_var[var_name].append(q)

    if not txn_variants:
        errors.append(f"'{SHEET_TXN}': no rows found — refusing to empty out txnVariants")
    if not comp_variants:
        errors.append(f"'{SHEET_COMP}': no rows found — refusing to empty out compVariants")

    for label, questions in (('appQuestions', app_q), ('infraQuestions', infra_q), ('logQuestions', log_q)):
        ids = [q['id'] for q in questions]
        dupes = {i for i in ids if ids.count(i) > 1}
        if dupes:
            errors.append(f'{label}: duplicate question ids {sorted(dupes)}')
    for label, variants in (('txnVariants', txn_variants), ('compVariants', comp_variants)):
        for variant, questions in variants.items():
            ids = [q['id'] for q in questions]
            dupes = {i for i in ids if ids.count(i) > 1}
            if dupes:
                errors.append(f'{label}.{variant}: duplicate question ids {sorted(dupes)}')

    return txn_variants, comp_variants, app_q, infra_q, log_q, errors


# ─── Serialize back to JS source, matching questions.js's existing style ──────

def _js_string(s):
    escaped = s.replace('\\', '\\\\').replace("'", "\\'").replace('\n', '\\n')
    return f"'{escaped}'"


def _serialize_question(q, indent):
    pad = '  ' * indent
    inner = '  ' * (indent + 1)
    opt_pad = '  ' * (indent + 2)
    lines = [pad + '{', inner + 'id: ' + _js_string(q['id']) + ',', inner + 'text: ' + _js_string(q['text']) + ',']
    if q.get('hint'):
        lines.append(inner + 'hint: ' + _js_string(q['hint']) + ',')
    lines.append(inner + 'options: [')
    opt_lines = [f"{opt_pad}{{ text: {_js_string(o['text'])}, score: {o['score']} }}" for o in q['options']]
    lines.append(',\n'.join(opt_lines))
    lines.append(inner + ']')
    lines.append(pad + '}')
    return '\n'.join(lines)


def _serialize_question_array(questions, indent):
    if not questions:
        return '[]'
    pad = '  ' * indent
    items = [_serialize_question(q, indent + 1) for q in questions]
    return '[\n' + ',\n'.join(items) + '\n' + pad + ']'


def _banner(key):
    return key.upper().replace('_', ' ')


_DASH = '\u2500'


def _serialize_variants(variants):
    pad_key = '  '
    parts = []
    for key, questions in variants.items():
        label = _banner(key)
        rule = _DASH * max(4, 70 - len(label))
        banner_line = f'{pad_key}// {_DASH * 3} {label} {rule}\n\n'
        arr = _serialize_question_array(questions, 1)
        parts.append(f'{banner_line}{pad_key}{key}: {arr}')
    return '{\n\n' + ',\n\n'.join(parts) + '\n}'


def serialize_var(name, value):
    if isinstance(value, dict):
        return f'var {name} = {_serialize_variants(value)};'
    return f'var {name} = {_serialize_question_array(value, 0)};'


# ─── Locate each `var NAME = ...;` statement's exact span in the file ──────────

def _blank_comments(source):
    return re.sub(r'//[^\n]*', lambda m: ' ' * len(m.group(0)), source)


def _locate_statement(blanked, var_name):
    m = re.search(r'var\s+' + re.escape(var_name) + r'\s*=\s*', blanked)
    if not m:
        raise ValueError(f'{var_name} not found in {QUESTIONS_JS}')
    _, end = _parse_value(blanked, m.end())
    i = end
    while i < len(blanked) and blanked[i] in ' \t\r\n':
        i += 1
    if i < len(blanked) and blanked[i] == ';':
        i += 1
    return m.start(), i


def splice(source, replacements):
    """replacements: {var_name: new_statement_text}. Finds each var's current
    span and replaces all of them in one pass (positions computed up front,
    since replacing one would shift offsets for the others)."""
    blanked = _blank_comments(source)
    spans = {name: _locate_statement(blanked, name) for name in replacements}
    ordered = sorted(spans.items(), key=lambda kv: kv[1][0])

    out = []
    cursor = 0
    for name, (start, end) in ordered:
        out.append(source[cursor:start])
        out.append(replacements[name])
        cursor = end
    out.append(source[cursor:])
    return ''.join(out)


# ─── Diff summary ───────────────────────────────────────────────────────────

def _counts(variants_or_list):
    if isinstance(variants_or_list, dict):
        return {k: len(v) for k, v in variants_or_list.items()}
    return {'*': len(variants_or_list)}


def _ids(variants_or_list):
    if isinstance(variants_or_list, dict):
        return {k: {q['id'] for q in v} for k, v in variants_or_list.items()}
    return {'*': {q['id'] for q in variants_or_list}}


def print_diff(label, before, after):
    before_ids, after_ids = _ids(before), _ids(after)
    keys = sorted(set(before_ids) | set(after_ids))
    for key in keys:
        b, a = before_ids.get(key, set()), after_ids.get(key, set())
        added, removed = a - b, b - a
        tag = f'{label}.{key}' if key != '*' else label
        if added or removed:
            print(f'  {tag}: {len(b)} -> {len(a)} questions'
                  f'{"  +" + str(sorted(added)) if added else ""}'
                  f'{"  -" + str(sorted(removed)) if removed else ""}')
        elif len(b) != len(a):
            print(f'  {tag}: {len(b)} -> {len(a)} questions')


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(path):
        print(f'File not found: {path}')
        sys.exit(1)

    txn_before, comp_before, app_before, infra_before, log_before = load_questions()
    txn_after, comp_after, app_after, infra_after, log_after, errors = read_workbook(path)

    if errors:
        print(f'{len(errors)} problem(s) found — {QUESTIONS_JS} was NOT modified:')
        for e in errors:
            print(f'  ! {e}')
        sys.exit(1)

    with open(QUESTIONS_JS, 'r', encoding='utf-8') as f:
        source = f.read()

    backup_path = QUESTIONS_JS + '.bak'
    with open(backup_path, 'w', encoding='utf-8') as f:
        f.write(source)

    new_source = splice(source, {
        'txnVariants': serialize_var('txnVariants', txn_after),
        'compVariants': serialize_var('compVariants', comp_after),
        'appQuestions': serialize_var('appQuestions', app_after),
        'infraQuestions': serialize_var('infraQuestions', infra_after),
        'logQuestions': serialize_var('logQuestions', log_after),
    })

    with open(QUESTIONS_JS, 'w', encoding='utf-8') as f:
        f.write(new_source)

    print(f'Wrote {QUESTIONS_JS} (backup at {backup_path})')
    print_diff('txnVariants', txn_before, txn_after)
    print_diff('compVariants', comp_before, comp_after)
    print_diff('appQuestions', app_before, app_after)
    print_diff('infraQuestions', infra_before, infra_after)
    print_diff('logQuestions', log_before, log_after)


if __name__ == '__main__':
    main()
