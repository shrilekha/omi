"""
Exports every OMI assessment question and its answer options (all sector/country
variants) from frontend/questions.js into a single Excel workbook for content review.

Usage:
    python scripts/export_questions_to_excel.py

Output:
    OMI_Questions_Review.xlsx (project root)
"""
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
QUESTIONS_JS = os.path.join(ROOT, 'frontend', 'questions.js')
OUTPUT_XLSX = os.path.join(ROOT, 'OMI_Questions_Review.xlsx')


# ─── Minimal JS object/array literal parser (strings, numbers, {}, []) ─────────

def _skip_ws(s, i):
    n = len(s)
    while i < n and s[i] in ' \t\r\n':
        i += 1
    return i


def _parse_string(s, i):
    quote = s[i]
    i += 1
    out = []
    mapping = {'n': '\n', 't': '\t', 'r': '\r', "'": "'", '"': '"', '\\': '\\'}
    while s[i] != quote:
        if s[i] == '\\':
            out.append(mapping.get(s[i + 1], s[i + 1]))
            i += 2
        else:
            out.append(s[i])
            i += 1
    return ''.join(out), i + 1


def _parse_number(s, i):
    start = i
    while i < len(s) and (s[i].isdigit() or s[i] in '.-'):
        i += 1
    numstr = s[start:i]
    return (float(numstr) if '.' in numstr else int(numstr)), i


def _parse_key(s, i):
    i = _skip_ws(s, i)
    if s[i] in ("'", '"'):
        return _parse_string(s, i)
    start = i
    while s[i] not in ':' and s[i] not in ' \t\r\n':
        i += 1
    return s[start:i], i


def _parse_value(s, i):
    i = _skip_ws(s, i)
    c = s[i]
    if c == '{':
        return _parse_object(s, i)
    if c == '[':
        return _parse_array(s, i)
    if c in ("'", '"'):
        return _parse_string(s, i)
    return _parse_number(s, i)


def _parse_array(s, i):
    i += 1
    arr = []
    i = _skip_ws(s, i)
    if s[i] == ']':
        return arr, i + 1
    while True:
        i = _skip_ws(s, i)
        val, i = _parse_value(s, i)
        arr.append(val)
        i = _skip_ws(s, i)
        if s[i] == ',':
            i = _skip_ws(s, i + 1)
            if s[i] == ']':
                break
            continue
        break
    i = _skip_ws(s, i)
    return arr, i + 1


def _parse_object(s, i):
    i += 1
    obj = {}
    i = _skip_ws(s, i)
    if s[i] == '}':
        return obj, i + 1
    while True:
        i = _skip_ws(s, i)
        key, i = _parse_key(s, i)
        i = _skip_ws(s, i)
        i += 1  # ':'
        i = _skip_ws(s, i)
        val, i = _parse_value(s, i)
        obj[key] = val
        i = _skip_ws(s, i)
        if s[i] == ',':
            i = _skip_ws(s, i + 1)
            if s[i] == '}':
                break
            continue
        break
    i = _skip_ws(s, i)
    return obj, i + 1


def extract_literal(source, var_name):
    """Find `var {var_name} = <literal>;` and return the parsed Python value."""
    m = re.search(r'var\s+' + re.escape(var_name) + r'\s*=\s*', source)
    if not m:
        raise ValueError(f'{var_name} not found')
    i = m.end()
    value, _ = _parse_value(source, i)
    return value


def load_questions():
    with open(QUESTIONS_JS, 'r', encoding='utf-8') as f:
        raw = f.read()
    no_comments = re.sub(r'//[^\n]*', '', raw)

    txn_variants = extract_literal(no_comments, 'txnVariants')
    comp_variants = extract_literal(no_comments, 'compVariants')
    app_questions = extract_literal(no_comments, 'appQuestions')
    infra_questions = extract_literal(no_comments, 'infraQuestions')
    log_questions = extract_literal(no_comments, 'logQuestions')

    return txn_variants, comp_variants, app_questions, infra_questions, log_questions


# ─── Excel build ────────────────────────────────────────────────────────────

HEADER = [
    'Domain', 'Sector / Country Variant', 'Question ID', 'Question Text', 'Hint',
    'Option 1 (score 1)', 'Option 2 (score 2)', 'Option 3 (score 3)',
    'Option 4 (score 4)', 'Option 5 (score 5)', 'Option 6 (Not sure / N-A)',
]

# Not part of questions.js — the assessment UI (frontend/index.html, renderSection())
# appends this option to every question at render time, excluded from scoring.
# Keep this in sync with that literal string if it's ever reworded.
NA_OPTION_TEXT = "Not sure / not applicable to my role — I don’t have visibility into this"

HEADER_FILL = PatternFill(start_color='0F1923', end_color='0F1923', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')


def question_row(domain, variant, q):
    opts = {o['score']: o['text'] for o in q['options']}
    return [
        domain, variant, q['id'], q['text'], q.get('hint', ''),
        opts.get(1, ''), opts.get(2, ''), opts.get(3, ''), opts.get(4, ''), opts.get(5, ''),
        NA_OPTION_TEXT,
    ]


def build_sheet(ws, rows):
    ws.append(HEADER)
    for col in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    widths = [26, 22, 10, 46, 40, 34, 34, 34, 34, 34, 34]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'


def main():
    txn_variants, comp_variants, app_q, infra_q, log_q = load_questions()

    wb = Workbook()

    # Sheet 1: Business & Transaction Observability (all sector/country variants)
    ws1 = wb.active
    ws1.title = 'Txn Observability'
    rows = []
    for variant, questions in txn_variants.items():
        for q in questions:
            rows.append(question_row('Business & Transaction Observability', variant, q))
    build_sheet(ws1, rows)

    # Sheet 2: Compliance & Audit Readiness (all sector/country variants)
    ws2 = wb.create_sheet('Compliance & Audit')
    rows = []
    for variant, questions in comp_variants.items():
        for q in questions:
            rows.append(question_row('Compliance & Audit Readiness', variant, q))
    build_sheet(ws2, rows)

    # Sheet 3: shared domains (Application Performance, Infrastructure, Log)
    ws3 = wb.create_sheet('App, Infra & Log (shared)')
    rows = []
    for q in app_q:
        rows.append(question_row('Application Performance', 'all sectors', q))
    for q in infra_q:
        rows.append(question_row('Infrastructure & Network', 'all sectors', q))
    for q in log_q:
        rows.append(question_row('Log Management & Data Lake', 'all sectors', q))
    build_sheet(ws3, rows)

    # Sheet 4: everything in one place, for a single-pass review
    ws4 = wb.create_sheet('All questions (combined)')
    rows = []
    for variant, questions in txn_variants.items():
        for q in questions:
            rows.append(question_row('Business & Transaction Observability', variant, q))
    for q in app_q:
        rows.append(question_row('Application Performance', 'all sectors', q))
    for q in infra_q:
        rows.append(question_row('Infrastructure & Network', 'all sectors', q))
    for q in log_q:
        rows.append(question_row('Log Management & Data Lake', 'all sectors', q))
    for variant, questions in comp_variants.items():
        for q in questions:
            rows.append(question_row('Compliance & Audit Readiness', variant, q))
    build_sheet(ws4, rows)

    wb.save(OUTPUT_XLSX)
    print(f'Wrote {len(rows)} rows (combined sheet) to {OUTPUT_XLSX}')


if __name__ == '__main__':
    main()
