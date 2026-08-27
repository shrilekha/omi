"""
Aggregation queries over the `submissions` table for the /admin/reports page —
sliceable, published, analyst-style reports on OMI's own respondent data. Not
synced with or dependent on omi-benchmarks; this is OMI's own data, OMI's own
report.

Hard rule, not a toggle: these queries never select email/first_name/
last_name/role. A report only ever sees aggregates.

No sample-size suppression (explicit product choice) — a figure is always
returned, however small the sample, but every figure carries its `n` so it's
never misread as more reliable than it is.
"""
import os
import sqlite3
import statistics

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ENV = os.getenv('ENV', 'development')
_PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

SCORE_COLUMNS = [
    ('overall_score', 'Overall'),
    ('txn_score', 'Business & Transaction Observability'),
    ('app_score', 'Application Performance'),
    ('infra_score', 'Infrastructure & Network'),
    ('log_score', 'Log Management & Data Lake'),
    ('comp_score', 'Compliance & Audit Readiness'),
]


def _get_conn():
    if ENV == 'production':
        import pymysql
        return pymysql.connect(
            host=os.getenv('DB_HOST', 'localhost'),
            user=os.getenv('DB_USER'),
            password=os.getenv('DB_PASS'),
            db=os.getenv('DB_NAME'),
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor,
        )
    else:
        db_path = os.path.join(_PROJECT_ROOT, os.getenv('SQLITE_PATH', 'local_dev.db'))
        conn = sqlite3.connect(os.path.abspath(db_path))
        conn.row_factory = sqlite3.Row
        return conn


def _placeholder():
    return '%s' if ENV == 'production' else '?'


def _where_clause(sector, country, revenue_band, date_from, date_to):
    ph = _placeholder()
    clauses, params = [], []
    if sector:
        clauses.append(f'sector = {ph}')
        params.append(sector)
    if country:
        clauses.append(f'country = {ph}')
        params.append(country)
    if revenue_band:
        clauses.append(f'revenue_band = {ph}')
        params.append(revenue_band)
    if date_from:
        clauses.append(f'submitted_at >= {ph}')
        params.append(date_from)
    if date_to:
        # date_to is a calendar date (e.g. from a <input type=date>) — make it
        # inclusive of the whole day rather than stopping at 00:00:00.
        clauses.append(f'submitted_at < {ph}')
        params.append(f'{date_to} 23:59:59.999999')
    sql = (' WHERE ' + ' AND '.join(clauses)) if clauses else ''
    return sql, params


def get_filter_options():
    """Distinct sector/country/revenue_band values that actually have at
    least one submission, for the filter dropdowns — there's no point
    offering a filter value with nothing behind it."""
    conn = _get_conn()
    try:
        options = {}
        for col in ('sector', 'country', 'revenue_band'):
            sql = f"SELECT DISTINCT {col} FROM submissions WHERE {col} IS NOT NULL AND {col} != '' ORDER BY {col}"
            if ENV == 'production':
                with conn.cursor() as cur:
                    cur.execute(sql)
                    rows = cur.fetchall()
                options[col] = [r[col] for r in rows]
            else:
                options[col] = [r[0] for r in conn.execute(sql)]
        return options
    finally:
        conn.close()


def get_report_data(sector=None, country=None, revenue_band=None, date_from=None, date_to=None):
    """Returns {n, scores: {col: {mean, median, label}}, maturity_bands: {band: count}}
    for the given filter combination (all filters optional and independently
    combinable). n is the sample size behind every figure here."""
    where_sql, params = _where_clause(sector, country, revenue_band, date_from, date_to)
    conn = _get_conn()
    try:
        cols = ', '.join(c for c, _ in SCORE_COLUMNS)
        sql = f'SELECT {cols}, maturity_band FROM submissions{where_sql}'
        if ENV == 'production':
            with conn.cursor() as cur:
                cur.execute(sql, params)
                rows = cur.fetchall()
        else:
            rows = [dict(r) for r in conn.execute(sql, params)]
    finally:
        conn.close()

    n = len(rows)
    scores = {}
    for col, label in SCORE_COLUMNS:
        values = [r[col] for r in rows if r[col] is not None]
        scores[col] = {
            'label': label,
            'n': len(values),
            'mean': round(statistics.mean(values), 1) if values else None,
            'median': round(statistics.median(values), 1) if values else None,
        }

    maturity_bands = {}
    for r in rows:
        band = r.get('maturity_band') or 'Unscored'
        maturity_bands[band] = maturity_bands.get(band, 0) + 1

    return {'n': n, 'scores': scores, 'maturity_bands': maturity_bands}


def build_report_workbook(data, filters):
    """A single-sheet spreadsheet of one filtered report — the offline
    counterpart to the /admin/reports dashboard, same numbers either way."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'

    header_fill = PatternFill(start_color='0F1923', end_color='0F1923', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    ws.append(['Filter', 'Value'])
    for col in range(1, 3):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).fill = header_fill
    for label, value in (
        ('Sector', filters.get('sector') or 'All'),
        ('Country', filters.get('country') or 'All'),
        ('Revenue band', filters.get('revenue_band') or 'All'),
        ('Date from', filters.get('date_from') or 'All time'),
        ('Date to', filters.get('date_to') or 'All time'),
        ('Sample size (n)', data['n']),
    ):
        ws.append([label, value])

    ws.append([])
    header_row = ws.max_row + 1
    ws.append(['Domain', 'Mean', 'Median', 'n'])
    for col in range(1, 5):
        ws.cell(row=header_row, column=col).font = header_font
        ws.cell(row=header_row, column=col).fill = header_fill
    for col, stats in data['scores'].items():
        ws.append([stats['label'], stats['mean'], stats['median'], stats['n']])

    ws.append([])
    header_row = ws.max_row + 1
    ws.append(['Maturity band', 'Count'])
    for col in range(1, 3):
        ws.cell(row=header_row, column=col).font = header_font
        ws.cell(row=header_row, column=col).fill = header_fill
    for band, count in data['maturity_bands'].items():
        ws.append([band, count])

    for col, width in enumerate([32, 14, 10, 8], start=1):
        ws.column_dimensions[chr(64 + col)].width = width

    return wb
