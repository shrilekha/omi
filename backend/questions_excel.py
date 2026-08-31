"""
Shared Excel <-> frontend/questions.js logic, used by both the CLI scripts
(scripts/export_questions_to_excel.py / scripts/import_questions_from_excel.py)
and the /admin/questions/export and /admin/questions/import routes in app.py —
one implementation, two ways to run it (terminal or the admin web page).

Only the 5 data literals (txnVariants, compVariants, appQuestions,
infraQuestions, logQuestions) are ever touched on import — everything else in
questions.js (the header comment block, sectorArchetypeMap, resolveArchetype(),
getSections()) is left byte-for-byte untouched, since that file has real logic
in it, not just data. Import is all-or-nothing: any bad row blocks the write.
"""
import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
QUESTIONS_JS = os.path.join(ROOT, 'frontend', 'questions.js')

SHEET_TXN = 'Txn Observability'
SHEET_COMP = 'Compliance & Audit'
SHEET_SHARED = 'App, Infra & Log (shared)'

DOMAIN_TO_SHARED_VAR = {
    'Application Performance': 'appQuestions',
    'Infrastructure & Network': 'infraQuestions',
    'Log Management & Data Lake': 'logQuestions',
}

HEADER = [
    'Domain', 'Sector / Country Variant', 'Question ID', 'Question Text', 'Hint',
    'Option 1 (score 1)', 'Option 2 (score 2)', 'Option 3 (score 3)',
    'Option 4 (score 4)', 'Option 5 (score 5)', 'Option 6 (Not sure / N-A)',
]

# Not part of questions.js — the assessment UI (frontend/index.html, renderSection())
# appends this option to every question at render time, excluded from scoring.
# Keep this in sync with that literal string if it's ever reworded.
NA_OPTION_TEXT = "Not sure / not applicable to my role — I don’t have visibility into this"

HEADER_FILL = PatternFill(start_color='0F1923', end_color='0F1923', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True)
WRAP = Alignment(wrap_text=True, vertical='top')


# ─── Minimal JS object/array literal parser (strings, numbers, {}, []) ─────────

def _skip_ws(s, i):
    n = len(s)
    while i < n and s[i] in ' \t\r\n':
        i += 1
    return i


def _parse_string(s, i):
    quote = s[i]
    i += 1
    out = []
    mapping = {'n': '\n', 't': '\t', 'r': '\r', "'": "'", '"': '"', '\\': '\\'}
    while s[i] != quote:
        if s[i] == '\\':
            out.append(mapping.get(s[i + 1], s[i + 1]))
            i += 2
        else:
            out.append(s[i])
            i += 1
    return ''.join(out), i + 1


def _parse_number(s, i):
    start = i
    while i < len(s) and (s[i].isdigit() or s[i] in '.-'):
        i += 1
    numstr = s[start:i]
    return (float(numstr) if '.' in numstr else int(numstr)), i


def _parse_key(s, i):
    i = _skip_ws(s, i)
    if s[i] in ("'", '"'):
        return _parse_string(s, i)
    start = i
    while s[i] not in ':' and s[i] not in ' \t\r\n':
        i += 1
    return s[start:i], i


def _parse_value(s, i):
    i = _skip_ws(s, i)
    c = s[i]
    if c == '{':
        return _parse_object(s, i)
    if c == '[':
        return _parse_array(s, i)
    if c in ("'", '"'):
        return _parse_string(s, i)
    return _parse_number(s, i)


def _parse_array(s, i):
    i += 1
    arr = []
    i = _skip_ws(s, i)
    if s[i] == ']':
        return arr, i + 1
    while True:
        i = _skip_ws(s, i)
        val, i = _parse_value(s, i)
        arr.append(val)
        i = _skip_ws(s, i)
        if s[i] == ',':
            i = _skip_ws(s, i + 1)
            if s[i] == ']':
                break
            continue
        break
    i = _skip_ws(s, i)
    return arr, i + 1


def _parse_object(s, i):
    i += 1
    obj = {}
    i = _skip_ws(s, i)
    if s[i] == '}':
        return obj, i + 1
    while True:
        i = _skip_ws(s, i)
        key, i = _parse_key(s, i)
        i = _skip_ws(s, i)
        i += 1  # ':'
        i = _skip_ws(s, i)
        val, i = _parse_value(s, i)
        obj[key] = val
        i = _skip_ws(s, i)
        if s[i] == ',':
            i = _skip_ws(s, i + 1)
            if s[i] == '}':
                break
            continue
        break
    i = _skip_ws(s, i)
    return obj, i + 1


def extract_literal(source, var_name):
    """Find `var {var_name} = <literal>;` and return the parsed Python value."""
    m = re.search(r'var\s+' + re.escape(var_name) + r'\s*=\s*', source)
    if not m:
        raise ValueError(f'{var_name} not found')
    i = m.end()
    value, _ = _parse_value(source, i)
    return value


def load_questions():
    """Returns (txn_variants, comp_variants, app_questions, infra_questions,
    log_questions) parsed live from the current frontend/questions.js."""
    with open(QUESTIONS_JS, 'r', encoding='utf-8') as f:
        raw = f.read()
    no_comments = re.sub(r'//[^\n]*', '', raw)

    txn_variants = extract_literal(no_comments, 'txnVariants')
    comp_variants = extract_literal(no_comments, 'compVariants')
    app_questions = extract_literal(no_comments, 'appQuestions')
    infra_questions = extract_literal(no_comments, 'infraQuestions')
    log_questions = extract_literal(no_comments, 'logQuestions')

    return txn_variants, comp_variants, app_questions, infra_questions, log_questions


# ─── Excel build (export) ───────────────────────────────────────────────────

def question_row(domain, variant, q):
    opts = {o['score']: o['text'] for o in q['options']}
    return [
        domain, variant, q['id'], q['text'], q.get('hint', ''),
        opts.get(1, ''), opts.get(2, ''), opts.get(3, ''), opts.get(4, ''), opts.get(5, ''),
        NA_OPTION_TEXT,
    ]


def build_sheet(ws, rows):
    ws.append(HEADER)
    for col in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    widths = [26, 22, 10, 46, 40, 34, 34, 34, 34, 34, 34]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = 'A2'


def build_workbook():
    """The full question set as a Workbook — one sheet per domain grouping.
    Deliberately no combined/"everything at once" sheet: every question
    exists in exactly one place, so there's no way to edit a copy that
    silently isn't the one read on import (a real "which tab is the real
    one?" data-loss risk an earlier version of this workbook had)."""
    txn_variants, comp_variants, app_q, infra_q, log_q = load_questions()

    wb = Workbook()

    ws1 = wb.active
    ws1.title = SHEET_TXN
    rows = []
    for variant, questions in txn_variants.items():
        for q in questions:
            rows.append(question_row('Business & Transaction Observability', variant, q))
    build_sheet(ws1, rows)

    ws2 = wb.create_sheet(SHEET_COMP)
    rows = []
    for variant, questions in comp_variants.items():
        for q in questions:
            rows.append(question_row('Compliance & Audit Readiness', variant, q))
    build_sheet(ws2, rows)

    ws3 = wb.create_sheet(SHEET_SHARED)
    rows = []
    for q in app_q:
        rows.append(question_row('Application Performance', 'all sectors', q))
    for q in infra_q:
        rows.append(question_row('Infrastructure & Network', 'all sectors', q))
    for q in log_q:
        rows.append(question_row('Log Management & Data Lake', 'all sectors', q))
    build_sheet(ws3, rows)

    return wb


# ─── Read rows out of an uploaded workbook (import) ─────────────────────────

def _row_dicts(ws):
    """Yield (row_number, {header: value}) for every non-empty data row."""
    header = [c.value for c in ws[1]]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v in (None, '') for v in row):
            continue
        yield row_idx, dict(zip(header, row))


def _question_from_row(row_idx, ref, d, errors):
    qid = str(d.get('Question ID') or '').strip()
    text = str(d.get('Question Text') or '').strip()
    hint = str(d.get('Hint') or '').strip()

    if not qid:
        errors.append(f'{ref} row {row_idx}: missing Question ID')
    if not text:
        errors.append(f'{ref} row {row_idx}: missing Question Text')

    options = []
    for score in range(1, 6):
        col = f'Option {score} (score {score})'
        val = str(d.get(col) or '').strip()
        if not val:
            errors.append(f'{ref} row {row_idx} ({qid or "?"}): missing Option {score}')
            continue
        options.append({'text': val, 'score': score})

    q = {'id': qid, 'text': text, 'options': options}
    if hint:
        q['hint'] = hint
    return q


def read_workbook(wb):
    """Takes an already-loaded openpyxl Workbook. Returns (txn_variants,
    comp_variants, app_q, infra_q, log_q, errors)."""
    errors = []
    for required in (SHEET_TXN, SHEET_COMP, SHEET_SHARED):
        if required not in wb.sheetnames:
            errors.append(f"required sheet '{required}' not found in workbook")
    if errors:
        return None, None, None, None, None, errors

    txn_variants = {}
    for row_idx, d in _row_dicts(wb[SHEET_TXN]):
        variant = str(d.get('Sector / Country Variant') or '').strip()
        if not variant:
            errors.append(f"'{SHEET_TXN}' row {row_idx}: missing Sector / Country Variant")
            continue
        q = _question_from_row(row_idx, f"'{SHEET_TXN}'", d, errors)
        txn_variants.setdefault(variant, []).append(q)

    comp_variants = {}
    for row_idx, d in _row_dicts(wb[SHEET_COMP]):
        variant = str(d.get('Sector / Country Variant') or '').strip()
        if not variant:
            errors.append(f"'{SHEET_COMP}' row {row_idx}: missing Sector / Country Variant")
            continue
        q = _question_from_row(row_idx, f"'{SHEET_COMP}'", d, errors)
        comp_variants.setdefault(variant, []).append(q)

    app_q, infra_q, log_q = [], [], []
    target_by_var = {'appQuestions': app_q, 'infraQuestions': infra_q, 'logQuestions': log_q}
    for row_idx, d in _row_dicts(wb[SHEET_SHARED]):
        domain = str(d.get('Domain') or '').strip()
        var_name = DOMAIN_TO_SHARED_VAR.get(domain)
        if not var_name:
            errors.append(f"'{SHEET_SHARED}' row {row_idx}: unrecognized Domain '{domain}'")
            continue
        q = _question_from_row(row_idx, f"'{SHEET_SHARED}'", d, errors)
        target_by_var[var_name].append(q)

    if not txn_variants:
        errors.append(f"'{SHEET_TXN}': no rows found — refusing to empty out txnVariants")
    if not comp_variants:
        errors.append(f"'{SHEET_COMP}': no rows found — refusing to empty out compVariants")

    for label, questions in (('appQuestions', app_q), ('infraQuestions', infra_q), ('logQuestions', log_q)):
        ids = [q['id'] for q in questions]
        dupes = {i for i in ids if ids.count(i) > 1}
        if dupes:
            errors.append(f'{label}: duplicate question ids {sorted(dupes)}')
    for label, variants in (('txnVariants', txn_variants), ('compVariants', comp_variants)):
        for variant, questions in variants.items():
            ids = [q['id'] for q in questions]
            dupes = {i for i in ids if ids.count(i) > 1}
            if dupes:
                errors.append(f'{label}.{variant}: duplicate question ids {sorted(dupes)}')

    return txn_variants, comp_variants, app_q, infra_q, log_q, errors


def read_workbook_file(path_or_file):
    return read_workbook(load_workbook(path_or_file, data_only=True))


# ─── Serialize back to JS source, matching questions.js's existing style ──────

def _js_string(s):
    escaped = s.replace('\\', '\\\\').replace("'", "\\'").replace('\n', '\\n')
    return f"'{escaped}'"


def _serialize_question(q, indent):
    pad = '  ' * indent
    inner = '  ' * (indent + 1)
    opt_pad = '  ' * (indent + 2)
    lines = [pad + '{', inner + 'id: ' + _js_string(q['id']) + ',', inner + 'text: ' + _js_string(q['text']) + ',']
    if q.get('hint'):
        lines.append(inner + 'hint: ' + _js_string(q['hint']) + ',')
    lines.append(inner + 'options: [')
    opt_lines = [f"{opt_pad}{{ text: {_js_string(o['text'])}, score: {o['score']} }}" for o in q['options']]
    lines.append(',\n'.join(opt_lines))
    lines.append(inner + ']')
    lines.append(pad + '}')
    return '\n'.join(lines)


def _serialize_question_array(questions, indent):
    if not questions:
        return '[]'
    pad = '  ' * indent
    items = [_serialize_question(q, indent + 1) for q in questions]
    return '[\n' + ',\n'.join(items) + '\n' + pad + ']'


def _banner(key):
    return key.upper().replace('_', ' ')


_DASH = '\u2500'


def _serialize_variants(variants):
    pad_key = '  '
    parts = []
    for key, questions in variants.items():
        label = _banner(key)
        rule = _DASH * max(4, 70 - len(label))
        banner_line = f'{pad_key}// {_DASH * 3} {label} {rule}\n\n'
        arr = _serialize_question_array(questions, 1)
        parts.append(f'{banner_line}{pad_key}{key}: {arr}')
    return '{\n\n' + ',\n\n'.join(parts) + '\n}'


def serialize_var(name, value):
    if isinstance(value, dict):
        return f'var {name} = {_serialize_variants(value)};'
    return f'var {name} = {_serialize_question_array(value, 0)};'


# ─── Locate each `var NAME = ...;` statement's exact span in the file ──────────

def _blank_comments(source):
    return re.sub(r'//[^\n]*', lambda m: ' ' * len(m.group(0)), source)


def _locate_statement(blanked, var_name):
    m = re.search(r'var\s+' + re.escape(var_name) + r'\s*=\s*', blanked)
    if not m:
        raise ValueError(f'{var_name} not found in {QUESTIONS_JS}')
    _, end = _parse_value(blanked, m.end())
    i = end
    while i < len(blanked) and blanked[i] in ' \t\r\n':
        i += 1
    if i < len(blanked) and blanked[i] == ';':
        i += 1
    return m.start(), i


def splice(source, replacements):
    """replacements: {var_name: new_statement_text}. Finds each var's current
    span and replaces all of them in one pass (positions computed up front,
    since replacing one would shift offsets for the others)."""
    blanked = _blank_comments(source)
    spans = {name: _locate_statement(blanked, name) for name in replacements}
    ordered = sorted(spans.items(), key=lambda kv: kv[1][0])

    out = []
    cursor = 0
    for name, (start, end) in ordered:
        out.append(source[cursor:start])
        out.append(replacements[name])
        cursor = end
    out.append(source[cursor:])
    return ''.join(out)


# ─── Diff summary ───────────────────────────────────────────────────────────

def _by_id(variants_or_list):
    if isinstance(variants_or_list, dict):
        return {k: {q['id']: q for q in v} for k, v in variants_or_list.items()}
    return {'*': {q['id']: q for q in variants_or_list}}


def diff_lines(label, before, after):
    """Reports added/removed ids AND ids whose content changed (text, hint,
    or options) — comparing only id-set membership would silently call an
    edited question's text/options a "no change", which is the single most
    common kind of edit a content team actually makes."""
    before_map, after_map = _by_id(before), _by_id(after)
    keys = sorted(set(before_map) | set(after_map))
    lines = []
    for key in keys:
        b, a = before_map.get(key, {}), after_map.get(key, {})
        added = sorted(set(a) - set(b))
        removed = sorted(set(b) - set(a))
        changed = sorted(qid for qid in (set(a) & set(b)) if a[qid] != b[qid])
        if added or removed or changed:
            tag = f'{label}.{key}' if key != '*' else label
            bits = f'{tag}: {len(b)} -> {len(a)} questions'
            if added:
                bits += f'  +{added}'
            if removed:
                bits += f'  -{removed}'
            if changed:
                bits += f'  ~{changed}'
            lines.append(bits)
    return lines


def apply_import(txn_after, comp_after, app_after, infra_after, log_after):
    """Writes the merged question set into frontend/questions.js (backing up
    the previous version to questions.js.bak first) and returns a list of
    human-readable diff lines against what was there before."""
    txn_before, comp_before, app_before, infra_before, log_before = load_questions()

    with open(QUESTIONS_JS, 'r', encoding='utf-8') as f:
        source = f.read()

    backup_path = QUESTIONS_JS + '.bak'
    with open(backup_path, 'w', encoding='utf-8') as f:
        f.write(source)

    new_source = splice(source, {
        'txnVariants': serialize_var('txnVariants', txn_after),
        'compVariants': serialize_var('compVariants', comp_after),
        'appQuestions': serialize_var('appQuestions', app_after),
        'infraQuestions': serialize_var('infraQuestions', infra_after),
        'logQuestions': serialize_var('logQuestions', log_after),
    })

    with open(QUESTIONS_JS, 'w', encoding='utf-8') as f:
        f.write(new_source)

    lines = []
    lines += diff_lines('txnVariants', txn_before, txn_after)
    lines += diff_lines('compVariants', comp_before, comp_after)
    lines += diff_lines('appQuestions', app_before, app_after)
    lines += diff_lines('infraQuestions', infra_before, infra_after)
    lines += diff_lines('logQuestions', log_before, log_after)
    return lines
